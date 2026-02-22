from __future__ import annotations

import argparse
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.base import Base
from app.core.auth import hash_password
from app.core.config import settings
from app.db.session import SessionLocal, engine
from app.models import (
    AdminUser,
    Apartment,
    ApartmentTariffSetting,
    ChargeMode,
    Invoice,
    InvoiceItem,
    InvoiceStatus,
    Meter,
    MeterReading,
    Tariff,
    Tenancy,
    Tenant,
    UnitType,
    UtilityPayment,
    UtilityType,
)

MONTHS_UA = {
    "Січень": 1,
    "Лютий": 2,
    "Березень": 3,
    "Квітень": 4,
    "Травень": 5,
    "Червень": 6,
    "Липень": 7,
    "Серпень": 8,
    "Вересень": 9,
    "Жовтень": 10,
    "Листопад": 11,
    "Грудень": 12,
}


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


def parse_period(sheet_name: str) -> tuple[int, int]:
    month_ua, year_text = sheet_name.split()
    return int(year_text), MONTHS_UA[month_ua]


def map_utility_type(service_name: str) -> UtilityType | None:
    text = service_name.lower()
    if "електро" in text:
        return UtilityType.electricity
    if "газ" in text:
        return UtilityType.gas
    if "водовідвед" in text:
        return UtilityType.sewage
    if "водопост" in text or "вода" in text:
        return UtilityType.water
    return None


def map_unit(unit_name: str, charge_mode: ChargeMode) -> UnitType:
    if isinstance(unit_name, UnitType):
        return unit_name
    text = str(unit_name or "").lower()
    if charge_mode == ChargeMode.fixed:
        return UnitType.month
    if "квт" in text or "kwh" in text:
        return UnitType.kwh
    return UnitType.m3


def import_data(
    tariffs_xlsx: Path,
    utility_xlsx: Path,
    apartment_code: str = "IVASIUKA-1",
    apartment_address: str = "Івасюка, квартира",
) -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        db.add(AdminUser(username=settings.admin_username, password_hash=hash_password(settings.admin_password)))
        apartment = Apartment(code=apartment_code, address=apartment_address)
        tenant = Tenant(full_name="Поточний орендар", phone=None, access_code="TENANT-IVASIUKA")
        db.add_all([apartment, tenant])
        db.flush()

        utilities_wb = load_workbook(utility_xlsx, data_only=True)
        periods = sorted(parse_period(name) for name in utilities_wb.sheetnames)
        start_date = date(periods[0][0], periods[0][1], 1)
        tenancy = Tenancy(apartment_id=apartment.id, tenant_id=tenant.id, start_date=start_date, end_date=None)
        db.add(tenancy)

        tariffs_ws = load_workbook(tariffs_xlsx, data_only=True)["Тарифи"]
        metered_services: set[str] = set()
        for row in range(3, tariffs_ws.max_row + 1):
            service_name = tariffs_ws.cell(row, 1).value
            if not service_name:
                continue
            service_name = str(service_name).strip()
            price = to_decimal(tariffs_ws.cell(row, 4).value)
            unit_name = str(tariffs_ws.cell(row, 5).value or "за міс").strip()
            effective_raw = tariffs_ws.cell(row, 7).value
            if price is None:
                continue

            if isinstance(effective_raw, datetime):
                effective_from = effective_raw.date()
            elif isinstance(effective_raw, date):
                effective_from = effective_raw
            else:
                effective_from = start_date

            charge_mode = ChargeMode.metered if "лічиль" in unit_name.lower() else ChargeMode.fixed
            utility_type = map_utility_type(service_name)
            tariff = Tariff(
                apartment_id=apartment.id,
                service_name=service_name,
                charge_mode=charge_mode,
                utility_type=utility_type,
                price_per_unit=price,
                unit_name=map_unit(unit_name, charge_mode),
                effective_from=effective_from,
            )
            db.add(tariff)
            setting = ApartmentTariffSetting(
                apartment_id=apartment.id,
                service_name=service_name,
                provider_company=str(tariffs_ws.cell(row, 2).value).strip() if tariffs_ws.cell(row, 2).value else None,
                cabinet_url=str(tariffs_ws.cell(row, 6).value).strip() if tariffs_ws.cell(row, 6).value else None,
                cabinet_login=None,
                cabinet_password_encrypted=None,
                last_tariff_check_at=effective_raw if isinstance(effective_raw, datetime) else None,
            )
            existing_setting = db.scalar(
                select(ApartmentTariffSetting).where(
                    ApartmentTariffSetting.apartment_id == apartment.id,
                    ApartmentTariffSetting.service_name == service_name,
                )
            )
            if existing_setting is None:
                db.add(setting)
            if charge_mode == ChargeMode.metered:
                metered_services.add(service_name)

        db.flush()

        first_sheet_name = sorted(utilities_wb.sheetnames, key=lambda n: parse_period(n))[0]
        first_sheet = utilities_wb[first_sheet_name]
        meter_by_service: dict[str, Meter] = {}
        for row in range(3, 200):
            service_name = first_sheet.cell(row, 1).value
            if not service_name:
                continue
            service_name = str(service_name).strip()
            if service_name not in metered_services:
                continue

            initial = to_decimal(first_sheet.cell(row, 2).value)
            if initial is None:
                continue
            utility_type = map_utility_type(service_name) or UtilityType.water
            meter = Meter(
                apartment_id=apartment.id,
                service_name=service_name,
                utility_type=utility_type,
                serial_number=None,
                initial_reading=initial,
                installed_at=start_date,
            )
            db.add(meter)
            db.flush()
            meter_by_service[service_name] = meter

        for sheet_name in sorted(utilities_wb.sheetnames, key=lambda n: parse_period(n)):
            ws = utilities_wb[sheet_name]
            year, month = parse_period(sheet_name)
            invoice_items: list[InvoiceItem] = []
            items_total = Decimal("0")

            for row in range(3, 200):
                service_name = ws.cell(row, 1).value
                if not service_name:
                    continue
                service_name = str(service_name).strip()
                if service_name.lower().startswith("компенсація"):
                    break

                amount = to_decimal(ws.cell(row, 6).value)
                if amount is None:
                    continue

                consumption = to_decimal(ws.cell(row, 4).value) or Decimal("1")
                unit_price = to_decimal(ws.cell(row, 5).value) or amount
                unit_name = "за міс"
                tariff = db.scalar(
                    select(Tariff)
                    .where(Tariff.service_name == service_name, Tariff.effective_from <= date(year, month, 1))
                    .order_by(Tariff.effective_from.desc())
                )
                if tariff is not None:
                    unit_name = tariff.unit_name

                invoice_items.append(
                    InvoiceItem(
                        service_name=service_name,
                        utility_type=map_utility_type(service_name),
                        unit_name=map_unit(unit_name, ChargeMode.fixed if unit_name == "за міс" else ChargeMode.metered),
                        consumption=consumption,
                        unit_price=unit_price,
                        amount=amount,
                    )
                )
                items_total += amount

                meter = meter_by_service.get(service_name)
                if meter is not None:
                    end_value = to_decimal(ws.cell(row, 3).value)
                    if end_value is not None:
                        db.add(MeterReading(meter_id=meter.id, year=year, month=month, value=end_value))

            payment = to_decimal(ws.cell(17, 6).value) or Decimal("0")
            carry_over = to_decimal(ws.cell(18, 6).value) or Decimal("0")
            closing_balance = to_decimal(ws.cell(19, 6).value)
            if closing_balance is None:
                closing_balance = carry_over + items_total - payment
            status = InvoiceStatus.paid if closing_balance <= 0 else InvoiceStatus.unpaid
            invoice = Invoice(
                apartment_id=apartment.id,
                tenant_id=tenant.id,
                year=year,
                month=month,
                status=status,
                total_amount=items_total,
                carry_over_debt=carry_over,
                utility_payment_received=payment,
                closing_balance=closing_balance,
            )
            db.add(invoice)
            db.flush()
            for item in invoice_items:
                item.invoice_id = invoice.id
                db.add(item)
            if payment > 0:
                db.add(
                    UtilityPayment(
                        apartment_id=apartment.id,
                        tenant_id=tenant.id,
                        invoice_id=invoice.id,
                        year=year,
                        month=month,
                        amount=payment,
                        paid_at=date(year, month, 1),
                        note="Imported from Excel",
                        confirmed=True,
                    )
                )

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import utility data from Excel files.")
    parser.add_argument(
        "--tariffs",
        default="/data/tariffs.xlsx",
        help="Path to tariffs xlsx file.",
    )
    parser.add_argument(
        "--utility",
        default="/data/utilities.xlsx",
        help="Path to utility history xlsx file.",
    )
    parser.add_argument(
        "--apartment-code",
        default="IVASIUKA-1",
        help="Apartment code for imported data.",
    )
    parser.add_argument(
        "--apartment-address",
        default="Івасюка, квартира",
        help="Apartment address for imported data.",
    )
    args = parser.parse_args()

    import_data(
        tariffs_xlsx=Path(args.tariffs),
        utility_xlsx=Path(args.utility),
        apartment_code=args.apartment_code,
        apartment_address=args.apartment_address,
    )
    print("Import completed.")
